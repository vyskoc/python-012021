import pandas as pd

liberec = pd.read_csv("zam_liberec.csv")
plzen = pd.read_csv("zam_plzeň.csv")
praha = pd.read_csv("zam_praha.csv")
platy = pd.read_csv("platy_2021_02.csv")
vykazy = pd.read_csv("vykazy.csv")
liberec["mesto"] = "Liberec"
plzen["mesto"] = "Plzeň"
praha["mesto"] = "Praha"
zamestnanci = pd.concat([liberec, plzen, praha], ignore_index=True)
zamestnanciSplaty = pd.merge(zamestnanci, platy, on="cislo_zamestnance", how="outer")
vykazyZamestnancu = pd.merge(zamestnanciSplaty, vykazy, left_on="cislo_zamestnance", right_on="emloyee_id")
poctyHodin = vykazyZamestnancu.groupby(["project", "mesto"])["hours"].sum()
poctyHodin.to_excel("poctyHodin.xlsx", index=True, merge_cells=False)
poctyHodinNactene = pd.read_excel("poctyHodin.xlsx")
print(poctyHodinNactene)

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
#print(rozvrh_hodin[0][1])
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
      bunka = ws1.cell(radek,sloupec)
      #print(rozvrh_hodin[radek-1][sloupec-1])
      bunka.value = rozvrh_hodin[radek-1][sloupec-1]
      sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")